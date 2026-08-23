"""
Excel Inspector
===============

Modern replacement for the old xlrd-based XLSX reader.

Features:
- Reads modern .xlsx/.xlsm Excel files
- Validates file paths and extensions
- Lists worksheets
- Selects a worksheet by name or index
- Reports worksheet dimensions
- Detects actually populated rows/columns
- Extracts headers
- Extracts individual rows
- Extracts individual columns
- Displays a table preview
- Detects Python data types
- Counts empty/non-empty cells
- Converts rows into dictionaries using the header row
- Exports worksheet contents to CSV
- Exports worksheet contents to JSON
- Command-line interface
- Read-only mode for efficient processing of large Excel files
- Clear error handling

Install dependency:

    pip install openpyxl

Examples:

    python excel_inspector.py sample.xlsx

    python excel_inspector.py sample.xlsx --sheet 0

    python excel_inspector.py sample.xlsx --sheet "Sheet1"

    python excel_inspector.py sample.xlsx --row 2

    python excel_inspector.py sample.xlsx --column A

    python excel_inspector.py sample.xlsx --preview 20

    python excel_inspector.py sample.xlsx --export-csv output.csv

    python excel_inspector.py sample.xlsx --export-json output.json
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


# ---------------------------------------------------------------------------
# Formatting utilities
# ---------------------------------------------------------------------------

def serialise_value(value: Any) -> Any:
    """
    Convert Excel/Python values into JSON-safe values.
    """
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()

    return value


def display_value(value: Any) -> str:
    """
    Convert a cell value into a readable string.
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    return str(value)


def python_type_name(value: Any) -> str:
    """
    Return a user-friendly Python datatype name.
    """
    if value is None:
        return "empty"

    return type(value).__name__


# ---------------------------------------------------------------------------
# Excel inspector
# ---------------------------------------------------------------------------

class ExcelInspector:
    """
    High-level interface for inspecting an Excel workbook.
    """

    SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

    def __init__(
        self,
        filename: str | Path,
        data_only: bool = True,
        read_only: bool = True,
    ) -> None:

        self.path = Path(filename).expanduser().resolve()

        self._validate_file()

        self.workbook = load_workbook(
            filename=self.path,
            read_only=read_only,
            data_only=data_only,
        )

        self.sheet = self.workbook.worksheets[0]

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def _validate_file(self) -> None:
        """
        Validate the requested workbook before opening it.
        """

        if not self.path.exists():
            raise FileNotFoundError(
                f"Excel file does not exist:\n{self.path}"
            )

        if not self.path.is_file():
            raise ValueError(
                f"Path is not a file:\n{self.path}"
            )

        extension = self.path.suffix.lower()

        if extension not in self.SUPPORTED_EXTENSIONS:
            if extension == ".xls":
                raise ValueError(
                    "Legacy .xls detected.\n"
                    "openpyxl reads modern Excel formats such as .xlsx.\n"
                    "Convert this workbook to .xlsx or use xlrd specifically "
                    "for the old .xls format."
                )

            raise ValueError(
                f"Unsupported file format: {extension}\n"
                f"Supported formats: "
                f"{', '.join(sorted(self.SUPPORTED_EXTENSIONS))}"
            )

    # ------------------------------------------------------------------
    # Workbook information
    # ------------------------------------------------------------------

    @property
    def sheet_names(self) -> list[str]:
        return self.workbook.sheetnames

    def list_sheets(self) -> None:
        """
        Print all worksheets.
        """

        print("\nWORKSHEETS")
        print("-" * 70)

        for index, sheet_name in enumerate(self.sheet_names):
            marker = " <-- selected" if sheet_name == self.sheet.title else ""

            print(
                f"[{index:>2}] "
                f"{sheet_name}"
                f"{marker}"
            )

    # ------------------------------------------------------------------
    # Sheet selection
    # ------------------------------------------------------------------

    def select_sheet(self, selector: str | int) -> None:
        """
        Select worksheet by numeric index or sheet name.
        """

        if isinstance(selector, int):
            index = selector

        elif str(selector).strip().lstrip("-").isdigit():
            index = int(selector)

        else:
            name = str(selector)

            if name not in self.workbook.sheetnames:
                raise ValueError(
                    f"Worksheet '{name}' does not exist.\n"
                    f"Available worksheets: "
                    f"{', '.join(self.workbook.sheetnames)}"
                )

            self.sheet = self.workbook[name]
            return

        if index < 0 or index >= len(self.workbook.worksheets):
            raise IndexError(
                f"Worksheet index {index} is invalid. "
                f"Valid indexes are 0-"
                f"{len(self.workbook.worksheets) - 1}."
            )

        self.sheet = self.workbook.worksheets[index]

    # ------------------------------------------------------------------
    # Data extraction
    # ------------------------------------------------------------------

    def all_rows(self) -> list[list[Any]]:
        """
        Return worksheet contents as a list of rows.
        """

        return [
            list(row)
            for row in self.sheet.iter_rows(values_only=True)
        ]

    def populated_bounds(self) -> tuple[int, int]:
        """
        Calculate actual populated rows and columns.

        This is useful because worksheet max_row/max_column may occasionally
        include cells that were previously populated or formatted.
        """

        last_row = 0
        last_column = 0

        for row_number, row in enumerate(
            self.sheet.iter_rows(values_only=True),
            start=1,
        ):
            row_has_data = False

            for column_number, value in enumerate(row, start=1):

                if value is not None:
                    row_has_data = True
                    last_column = max(last_column, column_number)

            if row_has_data:
                last_row = row_number

        return last_row, last_column

    def headers(self, header_row: int = 1) -> list[Any]:
        """
        Extract column headings from a particular row.

        Excel row numbers are 1-based.
        """

        self._validate_row_number(header_row)

        return [
            cell.value
            for cell in next(
                self.sheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                )
            )
        ]

    def get_row(self, row_number: int) -> list[Any]:
        """
        Extract one complete row.
        """

        self._validate_row_number(row_number)

        return [
            cell.value
            for cell in next(
                self.sheet.iter_rows(
                    min_row=row_number,
                    max_row=row_number,
                )
            )
        ]

    def get_column(
        self,
        column: int | str,
        include_empty: bool = True,
    ) -> list[Any]:
        """
        Extract one Excel column.

        column can be:

            1
            2
            3

        or:

            A
            B
            C
            AA
        """

        column_index = self._normalise_column(column)

        values = []

        for row in self.sheet.iter_rows(
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            value = row[0]

            if include_empty or value is not None:
                values.append(value)

        return values

    def get_cell(
        self,
        row: int,
        column: int | str,
    ) -> Any:
        """
        Extract one cell.
        """

        self._validate_row_number(row)

        column_index = self._normalise_column(column)

        for current_row in self.sheet.iter_rows(
            min_row=row,
            max_row=row,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            return current_row[0]

        return None

    # ------------------------------------------------------------------
    # Dictionary conversion
    # ------------------------------------------------------------------

    def records(
        self,
        header_row: int = 1,
    ) -> list[dict[str, Any]]:
        """
        Convert the worksheet to a list of dictionaries.

        Example:

        Name | Age | City
        Kai  | 21  | Winchester

        becomes:

        {
            "Name": "Kai",
            "Age": 21,
            "City": "Winchester"
        }
        """

        headers = self.headers(header_row)

        clean_headers = []

        for index, header in enumerate(headers, start=1):
            if header is None or str(header).strip() == "":
                clean_headers.append(
                    f"Column_{get_column_letter(index)}"
                )
            else:
                clean_headers.append(str(header))

        output = []

        for row in self.sheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ):
            if not any(value is not None for value in row):
                continue

            record = {}

            for index, header in enumerate(clean_headers):

                value = row[index] if index < len(row) else None

                record[header] = serialise_value(value)

            output.append(record)

        return output

    # ------------------------------------------------------------------
    # Statistics
    # ------------------------------------------------------------------

    def statistics(self) -> dict[str, Any]:
        """
        Calculate worksheet statistics.
        """

        populated_rows, populated_columns = self.populated_bounds()

        total_cells = populated_rows * populated_columns

        non_empty = 0

        type_counts: dict[str, int] = {}

        if populated_rows > 0 and populated_columns > 0:

            for row in self.sheet.iter_rows(
                min_row=1,
                max_row=populated_rows,
                min_col=1,
                max_col=populated_columns,
                values_only=True,
            ):
                for value in row:

                    if value is not None:
                        non_empty += 1

                        datatype = python_type_name(value)

                        type_counts[datatype] = (
                            type_counts.get(datatype, 0) + 1
                        )

        empty = total_cells - non_empty

        return {
            "worksheet": self.sheet.title,
            "reported_max_rows": self.sheet.max_row,
            "reported_max_columns": self.sheet.max_column,
            "populated_rows": populated_rows,
            "populated_columns": populated_columns,
            "total_cells_in_populated_range": total_cells,
            "non_empty_cells": non_empty,
            "empty_cells": empty,
            "data_types": type_counts,
        }

    # ------------------------------------------------------------------
    # Console reports
    # ------------------------------------------------------------------

    def print_summary(self) -> None:
        """
        Print workbook and worksheet information.
        """

        stats = self.statistics()

        print("\n" + "=" * 70)
        print("EXCEL WORKBOOK INSPECTOR")
        print("=" * 70)

        print(f"File              : {self.path.name}")
        print(f"Path              : {self.path}")
        print(f"File size         : {self.path.stat().st_size:,} bytes")
        print(f"Number of sheets  : {len(self.workbook.sheetnames)}")
        print(f"Selected sheet    : {self.sheet.title}")

        print("\nWORKSHEET DIMENSIONS")
        print("-" * 70)

        print(
            f"Reported rows     : "
            f"{stats['reported_max_rows']:,}"
        )

        print(
            f"Reported columns  : "
            f"{stats['reported_max_columns']:,}"
        )

        print(
            f"Populated rows    : "
            f"{stats['populated_rows']:,}"
        )

        print(
            f"Populated columns : "
            f"{stats['populated_columns']:,}"
        )

        print(
            f"Non-empty cells   : "
            f"{stats['non_empty_cells']:,}"
        )

        print(
            f"Empty cells       : "
            f"{stats['empty_cells']:,}"
        )

        print("\nDATA TYPES")
        print("-" * 70)

        if stats["data_types"]:

            for datatype, count in sorted(
                stats["data_types"].items(),
                key=lambda item: item[0],
            ):
                print(
                    f"{datatype:<20} "
                    f"{count:>10,}"
                )

        else:
            print("No populated cells.")

    def print_headers(
        self,
        header_row: int = 1,
    ) -> None:
        """
        Display all column headings.
        """

        headers = self.headers(header_row)

        print(
            f"\nCOLUMN HEADERS — ROW {header_row}"
        )
        print("-" * 70)

        for index, header in enumerate(headers, start=1):

            letter = get_column_letter(index)

            print(
                f"{letter:>4} "
                f"[{index:>3}] : "
                f"{display_value(header)}"
            )

    def print_row(self, row_number: int) -> None:
        """
        Print a selected row.
        """

        values = self.get_row(row_number)

        print(
            f"\nROW {row_number}"
        )
        print("-" * 70)

        for index, value in enumerate(values, start=1):

            column_letter = get_column_letter(index)

            print(
                f"{column_letter:>4} : "
                f"{display_value(value)} "
                f"[{python_type_name(value)}]"
            )

    def print_column(
        self,
        column: int | str,
    ) -> None:
        """
        Print a selected column.
        """

        column_index = self._normalise_column(column)
        column_letter = get_column_letter(column_index)

        values = self.get_column(
            column_index,
            include_empty=True,
        )

        print(
            f"\nCOLUMN {column_letter}"
        )
        print("-" * 70)

        for row_number, value in enumerate(values, start=1):

            print(
                f"{row_number:>6} : "
                f"{display_value(value)}"
            )

    def print_preview(
        self,
        rows: int = 10,
    ) -> None:
        """
        Display the first N worksheet rows as a formatted table.
        """

        if rows <= 0:
            return

        data = []

        for index, row in enumerate(
            self.sheet.iter_rows(values_only=True),
            start=1,
        ):
            data.append(
                [display_value(value) for value in row]
            )

            if index >= rows:
                break

        if not data:
            print("\nWorksheet is empty.")
            return

        column_count = max(len(row) for row in data)

        widths = []

        for column_index in range(column_count):

            values = []

            for row in data:
                if column_index < len(row):
                    values.append(row[column_index])

            max_width = max(
                [len(value) for value in values] + [1]
            )

            # Prevent giant console columns.
            widths.append(
                min(max(max_width, 5), 30)
            )

        print(
            f"\nPREVIEW — FIRST {len(data)} ROWS"
        )
        print("-" * 70)

        header = ["Row"]

        header.extend(
            get_column_letter(i)
            for i in range(1, column_count + 1)
        )

        row_header_width = max(
            4,
            len(str(len(data))),
        )

        print(
            f"{header[0]:>{row_header_width}} | ",
            end="",
        )

        for index, title in enumerate(header[1:]):
            print(
                f"{title:<{widths[index]}}",
                end=" | ",
            )

        print()

        print(
            "-" *
            (
                row_header_width
                + sum(widths)
                + 3 * column_count
                + 3
            )
        )

        for row_number, row in enumerate(data, start=1):

            print(
                f"{row_number:>{row_header_width}} | ",
                end="",
            )

            for column_index in range(column_count):

                value = (
                    row[column_index]
                    if column_index < len(row)
                    else ""
                )

                width = widths[column_index]

                if len(value) > width:
                    value = (
                        value[: max(0, width - 3)]
                        + "..."
                    )

                print(
                    f"{value:<{width}}",
                    end=" | ",
                )

            print()

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_csv(
        self,
        output_filename: str | Path,
    ) -> Path:
        """
        Export the currently selected worksheet to CSV.
        """

        destination = Path(output_filename).expanduser().resolve()

        with destination.open(
            "w",
            newline="",
            encoding="utf-8-sig",
        ) as file:

            writer = csv.writer(file)

            for row in self.sheet.iter_rows(values_only=True):

                writer.writerow(
                    [
                        serialise_value(value)
                        for value in row
                    ]
                )

        return destination

    def export_json(
        self,
        output_filename: str | Path,
        header_row: int = 1,
    ) -> Path:
        """
        Export worksheet records to JSON.
        """

        destination = Path(output_filename).expanduser().resolve()

        data = self.records(
            header_row=header_row
        )

        with destination.open(
            "w",
            encoding="utf-8",
        ) as file:

            json.dump(
                data,
                file,
                indent=4,
                ensure_ascii=False,
                default=str,
            )

        return destination

    # ------------------------------------------------------------------
    # Validation helpers
    # ------------------------------------------------------------------

    def _validate_row_number(
        self,
        row_number: int,
    ) -> None:

        if row_number < 1:
            raise ValueError(
                "Excel row numbers start at 1."
            )

        if row_number > self.sheet.max_row:
            raise IndexError(
                f"Row {row_number} exceeds the worksheet "
                f"maximum row {self.sheet.max_row}."
            )

    @staticmethod
    def _normalise_column(
        column: int | str,
    ) -> int:

        if isinstance(column, int):

            if column < 1:
                raise ValueError(
                    "Excel column numbers start at 1."
                )

            return column

        text = str(column).strip().upper()

        if text.isdigit():

            number = int(text)

            if number < 1:
                raise ValueError(
                    "Excel column numbers start at 1."
                )

            return number

        try:
            return column_index_from_string(text)

        except ValueError as exc:
            raise ValueError(
                f"Invalid Excel column: {column}"
            ) from exc

    # ------------------------------------------------------------------
    # Resource management
    # ------------------------------------------------------------------

    def close(self) -> None:
        """
        Explicitly close the workbook.
        """

        self.workbook.close()

    def __enter__(self) -> "ExcelInspector":
        return self

    def __exit__(
        self,
        exc_type,
        exc_value,
        traceback,
    ) -> None:
        self.close()


# ---------------------------------------------------------------------------
# Command-line interface
# ---------------------------------------------------------------------------

def build_argument_parser() -> argparse.ArgumentParser:

    parser = argparse.ArgumentParser(
        description=(
            "Inspect and extract information from modern Excel "
            ".xlsx workbooks."
        )
    )

    parser.add_argument(
        "file",
        help="Path to the Excel workbook.",
    )

    parser.add_argument(
        "--sheet",
        default="0",
        help=(
            "Worksheet name or zero-based worksheet index. "
            "Default: 0"
        ),
    )

    parser.add_argument(
        "--header-row",
        type=int,
        default=1,
        help=(
            "Excel row containing column headers. "
            "Default: 1"
        ),
    )

    parser.add_argument(
        "--preview",
        type=int,
        default=10,
        help=(
            "Number of rows to display in preview. "
            "Default: 10"
        ),
    )

    parser.add_argument(
        "--row",
        type=int,
        help="Extract and display a specific Excel row.",
    )

    parser.add_argument(
        "--column",
        help=(
            "Extract a specific column, for example "
            "A, B, AA, 1, 2."
        ),
    )

    parser.add_argument(
        "--cell",
        help=(
            "Extract a specific cell such as A1 or C15."
        ),
    )

    parser.add_argument(
        "--export-csv",
        help="Export selected worksheet to a CSV file.",
    )

    parser.add_argument(
        "--export-json",
        help=(
            "Export selected worksheet to JSON using "
            "the header row as property names."
        ),
    )

    parser.add_argument(
        "--formulas",
        action="store_true",
        help=(
            "Return Excel formulas instead of their cached "
            "calculated values."
        ),
    )

    return parser


# ---------------------------------------------------------------------------
# Cell reference parser
# ---------------------------------------------------------------------------

def parse_cell_reference(
    reference: str,
) -> tuple[int, str]:

    reference = reference.strip().upper()

    letters = ""
    digits = ""

    for character in reference:

        if character.isalpha():

            if digits:
                raise ValueError(
                    f"Invalid cell reference: {reference}"
                )

            letters += character

        elif character.isdigit():
            digits += character

        else:
            raise ValueError(
                f"Invalid cell reference: {reference}"
            )

    if not letters or not digits:
        raise ValueError(
            f"Invalid cell reference: {reference}"
        )

    row = int(digits)

    if row < 1:
        raise ValueError(
            "Excel row numbers start at 1."
        )

    column_index_from_string(letters)

    return row, letters


# ---------------------------------------------------------------------------
# Main program
# ---------------------------------------------------------------------------

def main() -> int:

    parser = build_argument_parser()

    args = parser.parse_args()

    try:

        with ExcelInspector(
            filename=args.file,
            data_only=not args.formulas,
            read_only=True,
        ) as inspector:

            inspector.select_sheet(args.sheet)

            inspector.print_summary()

            inspector.list_sheets()

            inspector.print_headers(
                header_row=args.header_row
            )

            inspector.print_preview(
                rows=args.preview
            )

            # ----------------------------------------------------------
            # Specific row
            # ----------------------------------------------------------

            if args.row is not None:

                inspector.print_row(
                    args.row
                )

            # ----------------------------------------------------------
            # Specific column
            # ----------------------------------------------------------

            if args.column is not None:

                inspector.print_column(
                    args.column
                )

            # ----------------------------------------------------------
            # Specific cell
            # ----------------------------------------------------------

            if args.cell:

                row_number, column = parse_cell_reference(
                    args.cell
                )

                value = inspector.get_cell(
                    row=row_number,
                    column=column,
                )

                print(
                    f"\nCELL {args.cell.upper()}"
                )

                print("-" * 70)

                print(
                    f"Value : {display_value(value)}"
                )

                print(
                    f"Type  : {python_type_name(value)}"
                )

            # ----------------------------------------------------------
            # CSV export
            # ----------------------------------------------------------

            if args.export_csv:

                destination = inspector.export_csv(
                    args.export_csv
                )

                print(
                    "\nCSV exported successfully:"
                )

                print(destination)

            # ----------------------------------------------------------
            # JSON export
            # ----------------------------------------------------------

            if args.export_json:

                destination = inspector.export_json(
                    args.export_json,
                    header_row=args.header_row,
                )

                print(
                    "\nJSON exported successfully:"
                )

                print(destination)

        return 0

    except FileNotFoundError as error:

        print(
            f"\nFILE ERROR\n{error}",
            file=sys.stderr,
        )

        return 1

    except PermissionError as error:

        print(
            f"\nPERMISSION ERROR\n{error}",
            file=sys.stderr,
        )

        return 2

    except (ValueError, IndexError) as error:

        print(
            f"\nINPUT ERROR\n{error}",
            file=sys.stderr,
        )

        return 3

    except Exception as error:

        print(
            f"\nUNEXPECTED ERROR\n"
            f"{type(error).__name__}: {error}",
            file=sys.stderr,
        )

        return 99


if __name__ == "__main__":
    raise SystemExit(main())
